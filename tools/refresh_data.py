# -*- coding: utf-8 -*-
"""Actualisation nocturne de toutes les donnees de la page.

Deux blocs independants, executes l'un apres l'autre. L'echec de l'un
n'empeche pas l'autre d'aboutir ; l'ecriture n'a lieu que si au moins un
bloc a produit quelque chose.

1. Prix des matieres premieres. Le script telecharge la feuille mensuelle
   des cours de la Banque mondiale (Commodity Markets Outlook, dite Pink
   Sheet), en tire les moyennes annuelles, trimestrielles et les trois
   derniers mois, reecrit data/prix-brut.json et relance mkprix.py.

2. Macroeconomie congolaise. Le script interroge l'API de la Banque
   mondiale pour la croissance, l'inflation et le PIB reel de la RDC,
   reecrit la partie realisee de data/macro-rdc.json, puis relance
   mkmacro.py, qui redessine egalement les figures de la banque centrale
   a partir de data/bcc.json.

Seules les zones balisees d'index.html sont remplacees :

    <!-- PRIX:KPI:START -->    ... <!-- PRIX:KPI:END -->
    <!-- PRIX:FIGS:START -->   ... <!-- PRIX:FIGS:END -->
    <!-- PRIX:TAB:START -->    ... <!-- PRIX:TAB:END -->
    <!-- PRIX:STAMP:START -->  ... <!-- PRIX:STAMP:END -->
    <!-- MACRO:FIGS:START -->  ... <!-- MACRO:FIGS:END -->
    <!-- MACRO:TAB:START -->   ... <!-- MACRO:TAB:END -->
    <!-- BCC:FIGS:START -->    ... <!-- BCC:FIGS:END -->
    <!-- BCC:KPI:START -->     ... <!-- BCC:KPI:END -->
    <!-- SERIES:START -->      ... <!-- SERIES:END -->

Le texte d'analyse n'est jamais regenere : il contient des chiffres commentes
qui relevent du jugement de l'auteur, pas d'une mise a jour mecanique. Les
projections, le tableau des agregats annuels et les releves de la banque
centrale n'ont pas d'interface programmable connue : ils vivent dans
data/macro-rdc.json et data/bcc.json, versionnes avec le depot, et il suffit
d'y modifier une valeur pour que la page entiere se redessine.

Usage :
    python3 tools/refresh_data.py            # actualise
    python3 tools/refresh_data.py --dry-run  # affiche sans ecrire
"""
import datetime as dt
import io
import json
import os
import re
import shutil
import subprocess
import sys
import urllib.request

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
TOOLS = os.path.join(ROOT, "tools")
DATA = os.path.join(ROOT, "data")
INDEX = os.path.join(ROOT, "index.html")
BRUT = os.path.join(DATA, "prix-brut.json")
MACRO = os.path.join(DATA, "macro-rdc.json")

DRY = "--dry-run" in sys.argv

# Adresses candidates du classeur mensuel. La Banque mondiale a deja deplace ce
# fichier par le passe ; on essaie les adresses connues l'une apres l'autre et on
# echoue proprement si aucune ne repond, plutot que d'ecrire des donnees fausses.
CANDIDATS = [
    "https://thedocs.worldbank.org/en/doc/5d903e848db1d1b83e0ec8f744e55570-0350012021"
    "/related/CMO-Historical-Data-Monthly.xlsx",
    "https://thedocs.worldbank.org/en/doc/18675f1d1639c7a34d463f59263ba0a2-0050012025"
    "/related/CMO-Historical-Data-Monthly.xlsx",
    "https://www.worldbank.org/content/dam/sites/prospects/doc/CMO-Historical-Data-Monthly.xlsx",
]

# Correspondance entre les cles internes et les libelles de la Banque mondiale.
# La recherche se fait par sous-chaine, en minuscules, sur la ligne d'en-tete.
FEUILLE_PRIX = "Monthly Prices"
FEUILLE_IND = "Monthly Indices"
MAP_PRIX = {
    "copper": "copper",
    "tin": "tin",
    "gold": "gold",
    "brent": "crude oil, brent",
    "cocoa": "cocoa",
    "arab": "coffee, arabica",
    "rob": "coffee, robusta",
    "palm": "palm oil",
    "rubber": "rubber, tsr20",
    "logs": "logs, cameroon",
    "zinc": "zinc",
}
MAP_IND = {"mm": "metals & minerals", "pm": "precious metals"}

# Interface programmable de la Banque mondiale, pour les series longues de la
# RDC. La croissance et l'inflation sont reprises telles quelles ; le PIB reel
# en dollars constants sert a reconstruire l'indice de volume, base 2015 = 100.
API_BM = "https://api.worldbank.org/v2/country/COD/indicator/%s?format=json&per_page=400"
IND_BM = {
    "croissance": "NY.GDP.MKTP.KD.ZG",
    "inflation": "FP.CPI.TOTL.ZG",
    "niveau": "NY.GDP.MKTP.KD",
}
BASE_INDICE = 2015

MOIS_EN = ["January", "February", "March", "April", "May", "June",
           "July", "August", "September", "October", "November", "December"]
MOIS_FR = ["janvier", "f&#233;vrier", "mars", "avril", "mai", "juin",
           "juillet", "ao&#251;t", "septembre", "octobre", "novembre", "d&#233;cembre"]
ABBR_EN = ["Jan", "Feb", "Mar", "Apr", "May", "Jun",
           "Jul", "Aug", "Sep", "Oct", "Nov", "Dec"]
ABBR_FR = ["janv.", "f&#233;vr.", "mars", "avr.", "mai", "juin",
           "juil.", "ao&#251;t", "sept.", "oct.", "nov.", "d&#233;c."]


# Garde-fou : la page ne doit contenir aucun numero de telephone personnel.
_TEL = re.compile(r"(?:tel:|\+\s?1[\s.\-]?)\(?\d{3}\)?[\s.\-]?\d{3}[\s.\-]?\d{4}")


def log(*a):
    print("[refresh]", *a, flush=True)


# --------------------------------------------------------------- telechargement
def telecharge():
    dernier = None
    for url in CANDIDATS:
        try:
            log("essai", url.rsplit("/", 1)[-1], "sur", url.split("/")[2])
            req = urllib.request.Request(url, headers={"User-Agent": "Mozilla/5.0"})
            with urllib.request.urlopen(req, timeout=90) as r:
                blob = r.read()
            if len(blob) < 50000 or blob[:2] != b"PK":
                raise ValueError("reponse trop courte ou pas un xlsx")
            log("recu", len(blob), "octets")
            return blob
        except Exception as e:                                   # noqa: BLE001
            dernier = e
            log("  echec :", e)
    raise SystemExit(
        "Aucune adresse du Pink Sheet n'a repondu. Derniere erreur : %s\n"
        "Verifiez l'adresse du classeur mensuel sur\n"
        "  https://www.worldbank.org/en/research/commodity-markets\n"
        "puis ajoutez-la en tete de CANDIDATS dans tools/refresh_data.py." % dernier)


# --------------------------------------------------------------- lecture xlsx
def periode(v):
    """'2026M07' ou datetime -> (annee, mois) ; None si illisible."""
    if isinstance(v, dt.datetime):
        return v.year, v.month
    s = str(v).strip()
    m = re.match(r"^(\d{4})\s*[Mm]\s*(\d{1,2})$", s)
    if m:
        return int(m.group(1)), int(m.group(2))
    m = re.match(r"^(\d{4})-(\d{1,2})", s)
    if m:
        return int(m.group(1)), int(m.group(2))
    return None


def lit_feuille(blob, nom, cibles):
    """Rend {cle: {(annee, mois): valeur}} pour la feuille demandee."""
    from openpyxl import load_workbook
    wb = load_workbook(io.BytesIO(blob), data_only=True, read_only=True)
    if nom not in wb.sheetnames:
        proches = [s for s in wb.sheetnames if nom.split()[-1].lower() in s.lower()]
        if not proches:
            raise SystemExit("feuille introuvable : %s (feuilles : %s)"
                             % (nom, wb.sheetnames))
        nom = proches[0]
    ws = wb[nom]
    grille = [list(r) for r in ws.iter_rows(max_row=12, values_only=True)]

    # ligne d'en-tete : celle ou l'on reconnait le plus de cibles
    besoin = {k: v.lower() for k, v in cibles.items()}
    meilleure, score_max = None, 0
    for r in grille:
        txt = [str(c).strip().lower() if c is not None else "" for c in r]
        sc = sum(1 for lab in besoin.values() if any(lab == t or lab in t for t in txt))
        if sc > score_max:
            meilleure, score_max = txt, sc
    if not meilleure or score_max < len(besoin) // 2:
        raise SystemExit("ligne d'en-tete introuvable dans la feuille %s" % nom)

    col = {}
    for cle, lab in besoin.items():
        idx = None
        for i, t in enumerate(meilleure):
            if t == lab:
                idx = i
                break
        if idx is None:
            for i, t in enumerate(meilleure):
                if lab in t:
                    idx = i
                    break
        if idx is None:
            log("  colonne absente :", cle, "/", lab)
        else:
            col[cle] = idx

    out = {k: {} for k in col}
    for r in ws.iter_rows(values_only=True):
        if not r or r[0] is None:
            continue
        p = periode(r[0])
        if not p:
            continue
        for cle, i in col.items():
            if i < len(r) and isinstance(r[i], (int, float)):
                out[cle][p] = float(r[i])
    return out


# --------------------------------------------------------------- agregation
def moy(serie, cles):
    v = [serie[k] for k in cles if k in serie]
    return round(sum(v) / len(v), 4) if v else None


def construit(series):
    """Rend (labels, valeurs) : 3 annees, 2 trimestres, 3 derniers mois."""
    toutes = set()
    for s in series.values():
        toutes |= set(s.keys())
        break
    for s in series.values():
        toutes &= set(s.keys())
    if not toutes:
        raise SystemExit("aucune periode commune a toutes les series")
    dernier = max(toutes)
    an, mo = dernier

    mois3 = sorted(toutes)[-3:]

    # trimestres complets de l'annee courante, les deux plus recents
    trim = []
    for q in (1, 2, 3, 4):
        cles = [(an, m) for m in range(3 * q - 2, 3 * q + 1)]
        if all(c in toutes for c in cles):
            trim.append((q, cles))
    trim = trim[-2:]

    annees = [a for a in range(an - 3, an) ]
    annees = [a for a in annees
              if all((a, m) in toutes for m in range(1, 13))][-3:]
    if not annees:
        annees = [an - 3, an - 2, an - 1]

    labels, agreg = [], []
    for a in annees:
        labels.append([str(a), str(a)])
        agreg.append([(a, m) for m in range(1, 13)])
    yy = str(an)[-2:]
    for q, cles in trim:
        labels.append(["Q%d-%s" % (q, yy), "T%d-%s" % (q, yy)])
        agreg.append(cles)
    for (a, m) in mois3:
        labels.append(["%s %s" % (ABBR_EN[m - 1], str(a)[-2:]),
                       "%s %s" % (ABBR_FR[m - 1], str(a)[-2:])])
        agreg.append([(a, m)])

    valeurs = {}
    for cle, s in series.items():
        col = [moy(s, c) for c in agreg]
        if any(v is None for v in col):
            log("  serie incomplete, conservee telle quelle :", cle)
            continue
        valeurs[cle] = [round(v, 4 if abs(v) < 100 else 1) for v in col]
    return labels, valeurs, dernier


LIBELLES = {
    "copper": ("Copper, $/t", "Cuivre, $/t"),
    "tin": ("Tin, $/t", "&#201;tain, $/t"),
    "zinc": ("Zinc, $/t", "Zinc, $/t"),
    "gold": ("Gold, $/toz", "Or, $/once"),
    "brent": ("Brent, $/bbl", "Brent, $/baril"),
    "cocoa": ("Cocoa, $/kg", "Cacao, $/kg"),
    "arab": ("Coffee arabica, $/kg", "Caf&#233; arabica, $/kg"),
    "rob": ("Coffee robusta, $/kg", "Caf&#233; robusta, $/kg"),
    "palm": ("Palm oil, $/t", "Huile de palme, $/t"),
    "rubber": ("Rubber TSR20, $/kg", "Caoutchouc TSR20, $/kg"),
    "logs": ("Logs Africa, $/cum", "Grumes Afrique, $/m3"),
    "mm": ("Metals and minerals, 2010=100", "M&#233;taux et min&#233;raux, 2010=100"),
    "pm": ("Precious metals, 2010=100", "M&#233;taux pr&#233;cieux, 2010=100"),
}


def arrondi(v):
    return None if v is None else round(v, 4 if abs(v) < 100 else 2)


def historique(series):
    """Rend l'historique complet : chaque mois, chaque trimestre, chaque annee.

    On ne tronque rien : la Banque mondiale publie les cours mensuels depuis
    1960, et les trois tableaux reprennent tout ce que le classeur contient.
    Une case reste vide quand la serie concernee ne couvre pas la periode.
    """
    cles = [k for k in LIBELLES if k in series]
    if not cles:
        return None
    periodes = sorted(set().union(*(set(series[k]) for k in cles)))
    if not periodes:
        return None

    mois = [["%04d-%02d" % p] + [arrondi(series[k].get(p)) for k in cles]
            for p in periodes]

    trim = []
    for a in range(periodes[0][0], periodes[-1][0] + 1):
        for q in (1, 2, 3, 4):
            fenetre = [(a, m) for m in range(3 * q - 2, 3 * q + 1)]
            if not any(p in periodes for p in fenetre):
                continue
            trim.append(["%dQ%d" % (a, q)]
                        + [arrondi(moy(series[k], fenetre)) for k in cles])

    an = []
    for a in range(periodes[0][0], periodes[-1][0] + 1):
        fenetre = [(a, m) for m in range(1, 13)]
        presents = [p for p in fenetre if p in periodes]
        if len(presents) < 12:          # annee incomplete : moyenne trompeuse
            continue
        an.append([a] + [arrondi(moy(series[k], fenetre)) for k in cles])

    log("historique complet : %d mois, %d trimestres, %d annees, %d series"
        % (len(mois), len(trim), len(an), len(cles)))
    return {"cles": cles,
            "cols_en": [LIBELLES[k][0] for k in cles],
            "cols_fr": [LIBELLES[k][1] for k in cles],
            "mois": mois, "trim": trim, "an": an}


# --------------------------------------------------------------- reecriture
def remplace(S, nom, contenu):
    a = S.index("<!-- %s:START -->" % nom) + len("<!-- %s:START -->" % nom)
    b = S.index("<!-- %s:END -->" % nom)
    return S[:a] + contenu + S[b:]


def stamp(dernier, aujourdhui):
    a, m = dernier
    return ('<p class="stamp">'
            '<span class="l-en">Latest World Bank monthly reading: %s %d &#183; '
            'series refreshed automatically on %d %s %d</span>'
            '<span class="l-fr" lang="fr">Derni&#232;re observation mensuelle de la Banque '
            'mondiale&#8239;: %s %d &#183; s&#233;ries actualis&#233;es automatiquement '
            'le %d %s %d</span></p>'
            % (MOIS_EN[m - 1], a,
               aujourdhui.day, MOIS_EN[aujourdhui.month - 1], aujourdhui.year,
               MOIS_FR[m - 1], a,
               aujourdhui.day, MOIS_FR[aujourdhui.month - 1], aujourdhui.year))


# --------------------------------------------------------- macro : Banque mondiale
def indicateur(code):
    """Rend {annee: valeur} pour un indicateur de la Banque mondiale."""
    req = urllib.request.Request(API_BM % code, headers={"User-Agent": "Mozilla/5.0"})
    with urllib.request.urlopen(req, timeout=60) as r:
        paquet = json.loads(r.read().decode("utf-8"))
    if not isinstance(paquet, list) or len(paquet) < 2 or not paquet[1]:
        raise ValueError("reponse vide pour %s" % code)
    out = {}
    for o in paquet[1]:
        if o.get("value") is None:
            continue
        out[int(o["date"])] = float(o["value"])
    if len(out) < 20:
        raise ValueError("serie trop courte pour %s : %d points" % (code, len(out)))
    return out


def macro_bm(m):
    """Actualise la partie realisee de macro-rdc.json. Rend le nombre d'annees."""
    brut = {}
    for cle, code in IND_BM.items():
        brut[cle] = indicateur(code)
        log("  %s : %d annees, jusqu'a %d"
            % (code, len(brut[cle]), max(brut[cle])))

    proj = int(m["proj"]["annee"])
    # Aucune troncature : on prend toutes les annees que la Banque mondiale
    # documente pour les trois indicateurs a la fois. Seule l'annee de
    # projection est ecartee : celle-ci reste sous la main de l'auteur.
    ans = sorted(a for a in brut["croissance"]
                 if a < proj
                 and a in brut["inflation"] and a in brut["niveau"])
    if len(ans) < 25:
        raise ValueError("trop peu d'annees communes : %d" % len(ans))
    if BASE_INDICE not in brut["niveau"]:
        raise ValueError("annee de base %d absente du PIB reel" % BASE_INDICE)

    base = brut["niveau"][BASE_INDICE]
    m["annees"] = ans
    m["croissance"] = [round(brut["croissance"][a], 1) for a in ans]
    m["inflation"] = [round(brut["inflation"][a], 1) for a in ans]
    m["indice"] = [round(100.0 * brut["niveau"][a] / base, 1) for a in ans]
    m["source"] = "World Bank, World Development Indicators"
    return len(ans)


def bloc_macro(S, reg):
    """Redessine les figures macro et BCC. Rend (S, reg, resume)."""
    m = json.load(open(MACRO, encoding="utf-8"))
    try:
        n = macro_bm(m)
        log("series longues actualisees :", n, "annees jusqu'a", m["annees"][-1])
        if not DRY:
            json.dump(m, open(MACRO, "w", encoding="utf-8"),
                      ensure_ascii=False, indent=1)
    except Exception as e:                                       # noqa: BLE001
        # L'API n'a pas repondu : on redessine quand meme a partir du fichier
        # versionne, pour que le reste du bloc reste coherent.
        log("api Banque mondiale indisponible (%s) ; on garde data/macro-rdc.json" % e)

    if DRY:
        return S, reg, "macro (simule)"

    r = subprocess.run([sys.executable, "mkmacro.py"], cwd=TOOLS,
                       capture_output=True, text=True)
    if r.returncode:
        raise SystemExit("mkmacro.py a echoue :\n" + r.stdout + r.stderr)
    log(r.stdout.strip())

    lire = lambda n: open(os.path.join(TOOLS, n), encoding="utf-8").read()
    S = remplace(S, "MACRO:FIGS", lire("macro_figs.html"))
    S = remplace(S, "MACRO:TAB", lire("macro_tab.html"))
    S = remplace(S, "BCC:FIGS", lire("bcc_figs.html"))
    S = remplace(S, "BCC:KPI", lire("bcc_kpi.html"))
    reg.update(json.load(open(os.path.join(TOOLS, "series_macro.json"),
                              encoding="utf-8")))
    for f in ("macro_figs.html", "macro_tab.html", "bcc_figs.html",
              "bcc_kpi.html", "series_macro.json"):
        try:
            os.remove(os.path.join(TOOLS, f))
        except OSError:
            pass
    return S, reg, "macro RDC et BCC"


def exporte_csv(reg):
    """Ecrit une copie plate de chaque serie, pour qui prefere le fichier au bouton."""
    dossier = os.path.join(DATA, "csv")
    os.makedirs(dossier, exist_ok=True)
    ent = re.compile(r"&#(\d+);")
    dec = lambda s: ent.sub(lambda m: chr(int(m.group(1))), str(s))
    for cle, v in reg.items():
        lignes = [",".join('"%s"' % dec(c).replace('"', '""') for c in v["cols"])]
        for r in v["rows"]:
            lignes.append(",".join(
                ('"%s"' % dec(c).replace('"', '""')) if isinstance(c, str) else str(c)
                for c in r))
        with open(os.path.join(dossier, cle + ".csv"), "w",
                  encoding="utf-8-sig", newline="") as f:
            f.write("\r\n".join(lignes) + "\r\n")
    log("csv ecrits :", len(reg))


def bloc_prix(S, reg, aujourdhui):
    """Actualise les prix des matieres premieres. Rend (S, reg, resume)."""
    blob = telecharge()

    log("lecture des prix")
    prix = lit_feuille(blob, FEUILLE_PRIX, MAP_PRIX)
    log("lecture des indices")
    ind = lit_feuille(blob, FEUILLE_IND, MAP_IND)
    prix.update(ind)

    labels, valeurs, dernier = construit(prix)
    log("derniere periode :", "%dM%02d" % dernier, "| series :", len(valeurs))

    hist = historique(prix)

    brut = json.load(open(BRUT, encoding="utf-8"))
    brut["labels"] = labels
    brut["series"].update(valeurs)
    if hist:
        brut["historique"] = hist
    brut["edition_en"] = "%s %d" % (MOIS_EN[dernier[1] - 1], dernier[0])
    brut["edition_fr"] = "%s %d" % (MOIS_FR[dernier[1] - 1], dernier[0])
    brut["refreshed"] = aujourdhui.isoformat()

    if DRY:
        print(json.dumps({"labels": labels,
                          "copper": valeurs.get("copper"),
                          "gold": valeurs.get("gold")},
                         ensure_ascii=False, indent=1))
        return S, reg, "prix (simule)"

    json.dump(brut, open(BRUT, "w", encoding="utf-8"), ensure_ascii=False, indent=1)
    log("data/prix-brut.json reecrit")

    # --- regeneration des figures et du tableau
    r = subprocess.run([sys.executable, "mkprix.py"], cwd=TOOLS,
                       capture_output=True, text=True)
    if r.returncode:
        raise SystemExit("mkprix.py a echoue :\n" + r.stdout + r.stderr)
    log(r.stdout.strip())

    lire = lambda n: open(os.path.join(TOOLS, n), encoding="utf-8").read()
    S = remplace(S, "PRIX:KPI", lire("prix_kpi.html"))
    S = remplace(S, "PRIX:FIGS", lire("prix_figs.html"))
    S = remplace(S, "PRIX:TAB", lire("prix_tab.html"))
    S = remplace(S, "PRIX:STAMP", stamp(dernier, aujourdhui))
    reg.update(json.load(open(os.path.join(TOOLS, "series_prix.json"),
                              encoding="utf-8")))
    for f in ("prix_figs.html", "prix_tab.html", "prix_kpi.html", "series_prix.json"):
        try:
            os.remove(os.path.join(TOOLS, f))
        except OSError:
            pass
    return S, reg, "prix des matieres premieres"


def main():
    aujourdhui = dt.date.today()
    S = open(INDEX, encoding="utf-8").read()
    m = re.search(r'<!-- SERIES:START -->.*?id="series-data">(.*?)</script>', S, re.S)
    if not m:
        raise SystemExit("registre des series introuvable dans index.html")
    reg = json.loads(m.group(1))

    faits, rates = [], []
    for nom, fonction in (("prix", lambda s, r: bloc_prix(s, r, aujourdhui)),
                          ("macro", bloc_macro)):
        log("--- bloc", nom)
        try:
            S, reg, resume = fonction(S, reg)
            faits.append(resume)
        except (Exception, SystemExit) as e:                     # noqa: BLE001
            rates.append(nom)
            log("bloc", nom, "abandonne :", str(e).split("\n")[0])

    if not faits:
        raise SystemExit("aucun bloc n'a pu etre actualise : " + ", ".join(rates))
    if rates:
        log("blocs en echec, page laissee intacte sur ces parties :", ", ".join(rates))

    if DRY:
        log("--dry-run : rien n'est ecrit |", " + ".join(faits))
        return

    S = remplace(S, "SERIES",
                 '<script type="application/json" id="series-data">'
                 + json.dumps(reg, ensure_ascii=False, separators=(",", ":"))
                 + "</script>")

    if _TEL.search(S):
        raise SystemExit("garde-fou : un numero de telephone apparait dans la page, "
                         "ecriture annulee")

    shutil.copyfile(INDEX, INDEX + ".bak")
    open(INDEX, "w", encoding="utf-8").write(S)
    log("index.html mis a jour :", " + ".join(faits))

    json.dump(reg, open(os.path.join(DATA, "series.json"), "w", encoding="utf-8"),
              ensure_ascii=False, separators=(",", ":"))
    exporte_csv(reg)
    log("termine")


if __name__ == "__main__":
    main()
